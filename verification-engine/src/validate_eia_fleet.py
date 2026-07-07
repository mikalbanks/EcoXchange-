"""EIA fleet re-validation harness — benchmark Engine A against federal actuals.

This is the parallel of ``validate_pvdaq.py`` but at fleet scale. The prior
yardstick came from the now-decommissioned TypeScript fleet-validation harness,
which ran its own hand-rolled Hay-Davies physics plus a ``trackingBoost``
multiplier off an uncommitted baseline — a *contaminated* reference, never a
target to beat, and never to be cited (it is preserved once, quarantined, in
``archive/fleet-validation-legacy-snapshot.*``). This harness runs the **real**
Engine A pipeline (``expected_ac_energy`` -> ``apply_losses``) so the result is
the first clean benchmark on real physics, stated in absolute terms.

Ground truth is free, public-domain federal data:

  - **USPVDB**   -> DC capacity, ``axis_type`` (drives §1.2 tracking), tilt
  - **EIA-860**  -> nameplate / operating year / lat-lon / tilt / azimuth
  - **EIA-923**  -> monthly net generation = the actuals we score against

The cohort filters below are ported from the TypeScript harness
(``ecoxchange-fleet-validation/src/parsers/joiner.ts`` ``DEFAULT_JOIN_OPTIONS``
and ``backtest/outlier-analysis.ts``) and encode the EIA caveats: USPVDB DC (not
EIA-860 AC); tracking from ``axis_type``; drop months/years before the first full
operating year; plausibility bounds on capacity factor; exclude storage hybrids;
tag high-curtailment balancing authorities as a separate cohort.

The standing run is **two-way**: Engine A on NASA POWER vs Engine A on NSRDB.
That isolates the irradiance-source upgrade.

NOTE FOR CLAUDE CODE / OPERATORS: the federal-file column names and the live
NASA POWER / NSRDB fetch signatures must be confirmed against the downloaded
revision before trusting a production run — the metric and cohort math is correct
regardless; only the parse/fetch mapping needs live verification (same caveat as
``validate_pvdaq.py``). Run ``scripts/download-data.sh`` first.
"""
from __future__ import annotations

from dataclasses import dataclass, field, asdict
from datetime import date
from typing import Callable, Dict, List, Optional, Sequence
import math
import os

import numpy as np
import pandas as pd

from .verification_engine.config import (
    SystemConfig, Location, ArrayConfig, LossAssumptions,
)
from .verification_engine.irradiance import CANON
from .verification_engine.modelchain import expected_ac_energy
from .verification_engine.losses import apply_losses


HOURS_PER_YEAR = 8760
DEFAULT_AC_DC_RATIO = 1.3

# States used as a proxy for high-curtailment balancing authorities. BA is not in
# the federal join, so state is the best available proxy: CA ~ CAISO and TX ~
# ERCOT-West carry the bulk of US solar curtailment. Tagged as a SEPARATE cohort
# so curtailment-driven under-generation never contaminates the clean number.
HIGH_CURTAILMENT_STATES = {"CA", "TX"}


# ---------------------------------------------------------------------------
# Federal records + join (ported from joiner.ts / types.ts)
# ---------------------------------------------------------------------------

@dataclass
class USPVDBRecord:
    uspvdb_id: str
    name: str
    state: str
    county: Optional[str]
    latitude: float
    longitude: float
    capacity_ac_mw: Optional[float]
    capacity_dc_mw: float
    panel_technology: str
    axis_type: str           # "Fixed" | "Single Axis Tracking" | "Dual Axis Tracking"
    commissioning_year: int
    eia_plant_id: Optional[str]


@dataclass
class EIA860Record:
    eia_plant_id: str
    name_eia: str
    capacity_mw_860: float
    technology: str
    prime_mover: str
    latitude_eia: Optional[float]
    longitude_eia: Optional[float]
    operating_year: Optional[int]
    azimuth_deg: Optional[float]
    tilt_deg: Optional[float]


@dataclass
class EIA923PlantTotals:
    eia_plant_id: str
    name_923: str
    annual_mwh: float
    monthly_mwh: List[float]    # length 12, Jan->Dec
    year: int
    is_storage_hybrid: bool = False   # SUN reported alongside batteries / non-PV


@dataclass
class JoinOptions:
    min_capacity_mw_dc: float = 1.0
    max_capacity_mw_dc: float = 20.0
    exclude_partial_year: bool = True
    min_actual_cf_pct: float = 10.0
    max_actual_cf_pct: float = 30.0
    exclude_storage_hybrid: bool = True


DEFAULT_JOIN_OPTIONS = JoinOptions()


@dataclass
class JoinedPlant:
    eia_plant_id: str
    name: str
    latitude: float
    longitude: float
    state: str
    capacity_dc_mw: float
    capacity_ac_mw: Optional[float]
    panel_technology: str
    axis_type: str
    commissioning_year: int
    tilt_deg: float
    azimuth_deg: float
    tilt_source: str
    azimuth_source: str
    actual_annual_mwh: float
    actual_monthly_mwh: List[float]
    production_year: int
    actual_capacity_factor_pct: float
    high_curtailment: bool

    @property
    def is_tracking(self) -> bool:
        return "tracking" in self.axis_type.lower()


def estimate_tilt_from_latitude(lat: float) -> float:
    """NREL rule-of-thumb fixed-tilt from latitude (mirrors geo.ts)."""
    return max(0.0, min(45.0, abs(lat) * 0.76 + 3.1))


def join_datasets(uspvdb: Sequence[USPVDBRecord],
                  eia860: Sequence[EIA860Record],
                  eia923: Sequence[EIA923PlantTotals],
                  options: JoinOptions = DEFAULT_JOIN_OPTIONS) -> List[JoinedPlant]:
    """Join USPVDB + EIA-860 + EIA-923 on plant id and apply cohort filters.

    EIA-923 production is the universe (it is the ground truth); USPVDB and
    EIA-860 augment capacity / geometry / provenance.
    """
    by_860 = {r.eia_plant_id: r for r in eia860}
    by_us = {r.eia_plant_id: r for r in uspvdb if r.eia_plant_id}

    out: List[JoinedPlant] = []
    for gen in eia923:
        if options.exclude_storage_hybrid and gen.is_storage_hybrid:
            continue
        if gen.annual_mwh <= 0:
            continue

        us = by_us.get(gen.eia_plant_id)
        e860 = by_860.get(gen.eia_plant_id)

        # Capacity: USPVDB DC > EIA-860 nameplate (AC) x 1.3 > skip.
        if us is not None:
            cap_dc = us.capacity_dc_mw
            cap_ac = us.capacity_ac_mw
        elif e860 is not None:
            cap_ac = e860.capacity_mw_860
            cap_dc = e860.capacity_mw_860 * DEFAULT_AC_DC_RATIO
        else:
            continue
        if cap_dc is None or not (options.min_capacity_mw_dc <= cap_dc <= options.max_capacity_mw_dc):
            continue

        lat = (us.latitude if us else None)
        lat = lat if lat is not None else (e860.latitude_eia if e860 else None)
        lon = (us.longitude if us else None)
        lon = lon if lon is not None else (e860.longitude_eia if e860 else None)
        if lat is None or lon is None:
            continue

        # Tilt + azimuth with provenance.
        if e860 is not None and e860.tilt_deg is not None:
            tilt, tilt_source = e860.tilt_deg, "eia860"
        else:
            tilt, tilt_source = estimate_tilt_from_latitude(lat), "estimated"
        if e860 is not None and e860.azimuth_deg is not None:
            azimuth, azimuth_source = e860.azimuth_deg, "eia860"
        else:
            azimuth, azimuth_source = 180.0, "default"

        commissioning_year = (
            (us.commissioning_year if us else 0)
            or (e860.operating_year if e860 and e860.operating_year else 0)
        )

        # Partial-year filter: a plant commissioned during (or after) the
        # production year has no clean full operating year — biases deviation high.
        if (options.exclude_partial_year and commissioning_year > 0
                and commissioning_year >= gen.year):
            continue

        cf = (gen.annual_mwh / (cap_dc * HOURS_PER_YEAR)) * 100.0
        if cf < options.min_actual_cf_pct or cf > options.max_actual_cf_pct:
            continue

        monthly = list(gen.monthly_mwh)
        if not monthly or all(m == 0 for m in monthly):
            monthly = [gen.annual_mwh / 12.0] * 12

        state = (us.state if us else "") or ""
        out.append(JoinedPlant(
            eia_plant_id=gen.eia_plant_id,
            name=(us.name if us else None) or (e860.name_eia if e860 else None) or gen.name_923 or "Unnamed",
            latitude=lat, longitude=lon, state=state,
            capacity_dc_mw=cap_dc, capacity_ac_mw=cap_ac,
            panel_technology=(us.panel_technology if us else "Crystalline Silicon"),
            axis_type=(us.axis_type if us else "Fixed"),
            commissioning_year=commissioning_year or gen.year,
            tilt_deg=tilt, azimuth_deg=azimuth,
            tilt_source=tilt_source, azimuth_source=azimuth_source,
            actual_annual_mwh=gen.annual_mwh,
            actual_monthly_mwh=monthly,
            production_year=gen.year,
            actual_capacity_factor_pct=cf,
            high_curtailment=state.upper() in HIGH_CURTAILMENT_STATES,
        ))
    return out


# ---------------------------------------------------------------------------
# Per-plant backtest with Engine A
# ---------------------------------------------------------------------------

@dataclass
class PlantResult:
    eia_plant_id: str
    name: str
    state: str
    axis_type: str
    is_tracking: bool
    high_curtailment: bool
    irradiance_source: str
    expected_mwh: float
    actual_mwh: float
    deviation_pct: float            # (expected/actual - 1) * 100; + = model over-predicts
    expected_cf_pct: float
    actual_cf_pct: float
    within_10pct: bool
    within_5pct: bool
    likely_cause: str = ""


def _tz_for_longitude(lon: float) -> str:
    """Fixed-offset tz from longitude (Etc/GMT signs are inverted).

    Annual energy totals are effectively tz-invariant, so a fixed offset is fine
    for fleet-scale annual deviation without a lat/lon -> Olson lookup dependency.
    """
    off = int(round(lon / 15.0))
    return f"Etc/GMT-{off}" if off >= 0 else f"Etc/GMT+{abs(off)}"


def build_plant_config(plant: JoinedPlant) -> SystemConfig:
    """Map a joined federal record onto an Engine A SystemConfig.

    Tracking is driven straight off USPVDB ``axis_type`` (§1.2). Dual-axis is
    modelled as single-axis (the closest Engine A mount) and tagged as such.
    """
    return SystemConfig(
        name=plant.name,
        location=Location(
            latitude=plant.latitude, longitude=plant.longitude,
            tz=_tz_for_longitude(plant.longitude),
        ),
        array=ArrayConfig(
            surface_tilt=plant.tilt_deg,
            surface_azimuth=plant.azimuth_deg,
            dc_capacity_kw=plant.capacity_dc_mw * 1000.0,
            ac_capacity_kw=(plant.capacity_ac_mw * 1000.0
                            if plant.capacity_ac_mw
                            else plant.capacity_dc_mw / DEFAULT_AC_DC_RATIO * 1000.0),
            tracking=plant.is_tracking,
        ),
        losses=LossAssumptions(),
        commission_date=date(plant.commissioning_year, 1, 1),
    )


def _assert_no_look_ahead(weather: pd.DataFrame) -> None:
    """Guard: the model input frame must carry only weather, never any actuals.

    A single stray EIA-923 generation column leaking into the model frame would
    turn the benchmark into circular self-validation. We assert the columns are a
    subset of the canonical weather channels.
    """
    leaked = set(weather.columns) - set(CANON)
    if leaked:
        raise AssertionError(
            f"Look-ahead leak: model weather frame carries non-weather columns {sorted(leaked)}"
        )


def model_annual_mwh(cfg: SystemConfig, weather: pd.DataFrame, production_year: int) -> float:
    """Run Engine A end-to-end for one site-year and return net annual MWh."""
    _assert_no_look_ahead(weather)
    gross = expected_ac_energy(cfg, weather)
    gross_total = float(gross.sum())
    as_of = date(production_year, 7, 1)
    net_kwh, _ = apply_losses(cfg, gross_total, as_of)
    return net_kwh / 1000.0


def backtest_plant(plant: JoinedPlant, weather: pd.DataFrame,
                   source_name: str) -> PlantResult:
    cfg = build_plant_config(plant)
    expected_mwh = model_annual_mwh(cfg, weather, plant.production_year)
    actual_mwh = plant.actual_annual_mwh
    deviation = (expected_mwh / actual_mwh - 1.0) * 100.0 if actual_mwh else float("nan")
    denom = plant.capacity_dc_mw * HOURS_PER_YEAR
    result = PlantResult(
        eia_plant_id=plant.eia_plant_id, name=plant.name, state=plant.state,
        axis_type=plant.axis_type, is_tracking=plant.is_tracking,
        high_curtailment=plant.high_curtailment, irradiance_source=source_name,
        expected_mwh=expected_mwh, actual_mwh=actual_mwh, deviation_pct=deviation,
        expected_cf_pct=(expected_mwh / denom * 100.0) if denom else float("nan"),
        actual_cf_pct=plant.actual_capacity_factor_pct,
        within_10pct=abs(deviation) <= 10.0,
        within_5pct=abs(deviation) <= 5.0,
    )
    result.likely_cause = infer_outlier_cause(result)
    return result


def infer_outlier_cause(r: PlantResult) -> str:
    """Plausible-cause string for plants outside ±15% (ported from outlier-analysis.ts).

    The legacy ``trackingBoostApplied`` heuristic is gone: Engine A models
    single-axis tracking with real geometry (§1.2), so an under-prediction on a
    tracker is a genuine signal, not a missing fudge factor.
    """
    dev = r.deviation_pct
    if dev > 15:
        if not r.is_tracking and r.actual_cf_pct < 12:
            return "Possible: heavy shading, suboptimal orientation, or partial-year operation"
        return "Possible: curtailment, equipment issues, or inaccurate specs in EIA data"
    if dev < -15:
        if r.high_curtailment:
            return "Likely: curtailment (high-curtailment BA) — see curtailed cohort"
        return ("Possible: conservative model assumptions, higher-than-average "
                "irradiance year, or EIA data includes non-PV generation")
    return "Within normal model tolerance"


# ---------------------------------------------------------------------------
# Aggregation + cohorts
# ---------------------------------------------------------------------------

def aggregate(results: Sequence[PlantResult]) -> Dict[str, float]:
    """Mean / median / MAD / %±10 / %±5 / P25 / P75 over a cohort."""
    devs = np.array([r.deviation_pct for r in results
                     if r.deviation_pct == r.deviation_pct], dtype=float)
    n = int(len(devs))
    if n == 0:
        return {"n": 0}
    return {
        "n": n,
        "mean_deviation_pct": round(float(np.mean(devs)), 2),
        "median_deviation_pct": round(float(np.median(devs)), 2),
        "mean_abs_deviation_pct": round(float(np.mean(np.abs(devs))), 2),
        "p25_deviation_pct": round(float(np.percentile(devs, 25)), 2),
        "p75_deviation_pct": round(float(np.percentile(devs, 75)), 2),
        "pct_within_10": round(100.0 * float(np.mean(np.abs(devs) <= 10.0)), 1),
        "pct_within_5": round(100.0 * float(np.mean(np.abs(devs) <= 5.0)), 1),
    }


def split_cohorts(results: Sequence[PlantResult]) -> Dict[str, List[PlantResult]]:
    """Clean fixed-tilt (non-curtailed) / tracking / curtailed cohorts.

    Acceptance (§1.3) is read off ``clean_fixed`` and ``tracking``; the
    ``curtailed`` cohort is reported separately so its residual bias (a real,
    attributable effect) never moves the headline number.
    """
    clean_fixed, tracking, curtailed = [], [], []
    for r in results:
        if r.high_curtailment:
            curtailed.append(r)
        elif r.is_tracking:
            tracking.append(r)
        else:
            clean_fixed.append(r)
    return {"clean_fixed": clean_fixed, "tracking": tracking, "curtailed": curtailed}


def cohort_report(results: Sequence[PlantResult]) -> Dict[str, Dict[str, float]]:
    cohorts = split_cohorts(results)
    return {"all": aggregate(results), **{k: aggregate(v) for k, v in cohorts.items()}}


# ---------------------------------------------------------------------------
# Two-way driver: Engine A on NASA POWER vs Engine A on NSRDB
# ---------------------------------------------------------------------------

WeatherFetcher = Callable[[JoinedPlant], pd.DataFrame]


def run_fleet(plants: Sequence[JoinedPlant], fetch_weather: WeatherFetcher,
              source_name: str) -> List[PlantResult]:
    """Backtest every plant against one irradiance source; skip fetch failures."""
    results: List[PlantResult] = []
    for plant in plants:
        try:
            weather = fetch_weather(plant)
            results.append(backtest_plant(plant, weather, source_name))
        except Exception as exc:  # noqa: BLE001 — one bad site must not abort the run
            print(f"[warn] {plant.name} ({plant.eia_plant_id}) [{source_name}] failed: {exc}")
    return results


def nasa_power_fetcher(year: int) -> WeatherFetcher:
    from .verification_engine.irradiance import fetch_nasa_power

    def fetch(plant: JoinedPlant) -> pd.DataFrame:
        return fetch_nasa_power(
            build_plant_config(plant).location, f"{year}-01-01", f"{year}-12-31")
    return fetch


def nsrdb_fetcher(year: int) -> WeatherFetcher:
    from .verification_engine.irradiance import fetch_nsrdb
    api_key = os.environ.get("NREL_API_KEY")
    email = os.environ.get("NREL_EMAIL", "")
    if not api_key:
        raise RuntimeError("NSRDB run needs NREL_API_KEY (free at developer.nrel.gov/signup).")

    def fetch(plant: JoinedPlant) -> pd.DataFrame:
        return fetch_nsrdb(build_plant_config(plant).location, year, api_key, email)
    return fetch


def two_way_report(plants: Sequence[JoinedPlant], year: int) -> dict:
    """Standing comparison: A-on-NASA-POWER vs A-on-NSRDB (isolates irradiance)."""
    nasa = run_fleet(plants, nasa_power_fetcher(year), "nasa_power")
    out = {"year": year, "n_plants": len(plants),
           "nasa_power": cohort_report(nasa)}
    try:
        nsrdb = run_fleet(plants, nsrdb_fetcher(year), "nsrdb")
        out["nsrdb"] = cohort_report(nsrdb)
    except RuntimeError as exc:
        out["nsrdb"] = {"skipped": str(exc)}
    return out


# ---------------------------------------------------------------------------
# Federal-file parsers (field mappings ported from parsers/*.ts)
#
# These read the files dropped by scripts/download-data.sh. Column names follow
# the cited revisions; confirm against the downloaded revision before a
# production run (EIA workbook headers and USPVDB column names rotate annually).
# ---------------------------------------------------------------------------

def _num(v) -> Optional[float]:
    if v is None or (isinstance(v, float) and math.isnan(v)) or v == "":
        return None
    try:
        n = float(v)
    except (TypeError, ValueError):
        return None
    return n if math.isfinite(n) else None


def _normalize_tech(raw: str) -> str:
    v = (raw or "").lower()
    if "thin" in v:
        return "Thin Film"
    if "crystal" in v or "c-si" in v or "silicon" in v:
        return "Crystalline Silicon"
    return "Unknown"


def _normalize_axis(raw: str) -> str:
    v = (raw or "").lower()
    if "dual" in v:
        return "Dual Axis Tracking"
    if "single" in v or "tracking" in v:
        return "Single Axis Tracking"
    return "Fixed"


def parse_uspvdb(csv_path: str) -> List[USPVDBRecord]:
    """USPVDB CSV -> records (mirrors parsers/uspvdb.ts column mapping)."""
    df = pd.read_csv(csv_path, dtype=str, keep_default_na=False)
    out: List[USPVDBRecord] = []
    for r in df.to_dict("records"):
        lat, lon, cap_dc = _num(r.get("ylat")), _num(r.get("xlong")), _num(r.get("p_cap_dc"))
        if lat is None or lon is None or cap_dc is None:
            continue
        eia_id = (r.get("eia_id") or "").strip()
        out.append(USPVDBRecord(
            uspvdb_id=r.get("case_id") or r.get("objectid") or "",
            name=r.get("p_name") or "Unnamed", state=r.get("p_state") or "",
            county=r.get("p_county") or None, latitude=lat, longitude=lon,
            capacity_ac_mw=_num(r.get("p_cap_ac")), capacity_dc_mw=cap_dc,
            panel_technology=_normalize_tech(r.get("p_tech_primary") or ""),
            axis_type=_normalize_axis(r.get("p_axis") or ""),
            commissioning_year=int(_num(r.get("p_year")) or 0),
            eia_plant_id=eia_id if eia_id and eia_id != "0" else None,
        ))
    return out


def _clean_header(cell) -> str:
    """Collapse internal whitespace: 2024 EIA workbooks embed newlines in
    headers ("Reported\\nFuel Type Code"), older revisions use spaces."""
    import re
    return re.sub(r"\s+", " ", str(cell or "")).strip()


def _read_eia_grid(xlsx_path: str, header_token: str, sheet_match: Optional[str] = None):
    """Read an EIA workbook into (headers, rows), skipping its banner rows."""
    import openpyxl  # optional dep; only needed for live runs
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    name = next((n for n in wb.sheetnames if sheet_match and sheet_match.lower() in n.lower()),
                wb.sheetnames[0])
    grid = [list(row) for row in wb[name].iter_rows(values_only=True)]
    header_row = next((i for i in range(min(len(grid), 6))
                       if any(_clean_header(c) == header_token for c in grid[i])), 0)
    headers = [_clean_header(c) for c in grid[header_row]]
    return headers, grid[header_row + 1:]


def parse_eia860(xlsx_path: str) -> List[EIA860Record]:
    """EIA-860 solar workbook -> plant-level records (mirrors parsers/eia860.ts)."""
    headers, rows = _read_eia_grid(xlsx_path, "Plant Code", sheet_match="operable")
    idx = {h.lower(): i for i, h in enumerate(headers)}

    def col(name):
        return idx.get(name.lower(), -1)

    i_code, i_name, i_cap = col("Plant Code"), col("Plant Name"), col("Nameplate Capacity (MW)")
    i_tech, i_prime = col("Technology"), col("Prime Mover")
    i_lat, i_lon, i_op = col("Latitude"), col("Longitude"), col("Operating Year")
    i_az, i_tilt = col("Azimuth Angle"), col("Tilt Angle")
    if i_code < 0 or i_cap < 0:
        raise ValueError("EIA-860: required columns missing")

    by_plant: Dict[str, EIA860Record] = {}
    for row in rows:
        def g(i):
            return row[i] if 0 <= i < len(row) else None
        code = g(i_code)
        if code in (None, ""):
            continue
        tech, prime = str(g(i_tech) or ""), str(g(i_prime) or "")
        if "photovoltaic" not in tech.lower() and prime != "PV":
            continue
        pid = str(int(code)) if isinstance(code, float) else str(code)
        cap = _num(g(i_cap)) or 0.0
        if pid in by_plant:
            rec = by_plant[pid]
            rec.capacity_mw_860 += cap
            if rec.tilt_deg is None:
                rec.tilt_deg = _num(g(i_tilt))
            if rec.azimuth_deg is None:
                rec.azimuth_deg = _num(g(i_az))
        else:
            by_plant[pid] = EIA860Record(
                eia_plant_id=pid, name_eia=str(g(i_name) or ""), capacity_mw_860=cap,
                technology=tech, prime_mover=prime,
                latitude_eia=_num(g(i_lat)), longitude_eia=_num(g(i_lon)),
                operating_year=int(_num(g(i_op))) if _num(g(i_op)) else None,
                azimuth_deg=_num(g(i_az)) if i_az >= 0 else None,
                tilt_deg=_num(g(i_tilt)) if i_tilt >= 0 else None,
            )
    return list(by_plant.values())


_MONTHS = ["January", "February", "March", "April", "May", "June",
           "July", "August", "September", "October", "November", "December"]


def parse_eia923(xlsx_path: str, year: int) -> List[EIA923PlantTotals]:
    """EIA-923 monthly schedules -> per-plant solar (SUN) net generation.

    Sums generator rows to plant level. A plant carrying SUN alongside a battery
    or other prime mover is tagged ``is_storage_hybrid`` so the join can drop it.
    """
    headers, rows = _read_eia_grid(xlsx_path, "Plant Id", sheet_match="generation and fuel")
    idx = {h.lower(): i for i, h in enumerate(headers)}

    def col(name):
        return idx.get(name.lower(), -1)

    i_id, i_name = col("Plant Id"), col("Plant Name")
    i_fuel, i_prime = col("Reported Fuel Type Code"), col("Reported Prime Mover")
    i_months = [col(f"Netgen\n{m}") if col(f"Netgen\n{m}") >= 0 else col(f"Netgen {m}")
                for m in _MONTHS]

    agg: Dict[str, EIA923PlantTotals] = {}
    non_solar_prime: Dict[str, bool] = {}
    for row in rows:
        def g(i):
            return row[i] if 0 <= i < len(row) else None
        pid_raw = g(i_id)
        if pid_raw in (None, ""):
            continue
        pid = str(int(pid_raw)) if isinstance(pid_raw, float) else str(pid_raw)
        fuel, prime = str(g(i_fuel) or "").upper(), str(g(i_prime) or "").upper()
        if fuel != "SUN" or prime != "PV":
            # Track non-PV generation on the same plant id -> storage/hybrid flag.
            if pid in agg or fuel == "SUN":
                non_solar_prime[pid] = True
            continue
        monthly = [max(0.0, _num(g(i)) or 0.0) for i in i_months]
        if pid in agg:
            rec = agg[pid]
            rec.monthly_mwh = [a + b for a, b in zip(rec.monthly_mwh, monthly)]
            rec.annual_mwh = sum(rec.monthly_mwh)
        else:
            agg[pid] = EIA923PlantTotals(
                eia_plant_id=pid, name_923=str(g(i_name) or ""),
                annual_mwh=sum(monthly), monthly_mwh=monthly, year=year)
    for pid, rec in agg.items():
        rec.is_storage_hybrid = non_solar_prime.get(pid, False)
    return list(agg.values())


def load_and_join(data_dir: str, year: int,
                  options: JoinOptions = DEFAULT_JOIN_OPTIONS) -> List[JoinedPlant]:
    """Parse the three federal files under ``data_dir`` and join them."""
    uspvdb = parse_uspvdb(os.path.join(data_dir, "uspvdb", "uspvdb_centroids.csv"))
    eia860 = parse_eia860(os.path.join(data_dir, "eia860", "eia860_solar.xlsx"))
    eia923 = parse_eia923(os.path.join(data_dir, "eia923", "EIA923_Schedules.xlsx"), year)
    return join_datasets(uspvdb, eia860, eia923, options)


def main():
    import argparse
    import json

    ap = argparse.ArgumentParser(description="EIA fleet re-validation (Engine A).")
    ap.add_argument("--data-dir", default=os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "fleet"))
    ap.add_argument("--year", type=int, required=True)
    ap.add_argument("--out", default="eia_fleet_report.json")
    args = ap.parse_args()

    plants = load_and_join(args.data_dir, args.year)
    print(f"[info] Joined cohort: {len(plants)} plants in the 1-20 MW band")
    report = two_way_report(plants, args.year)
    with open(args.out, "w") as fh:
        json.dump(report, fh, indent=2)
    nasa_clean = report["nasa_power"].get("clean_fixed", {})
    print(f"[done] Wrote {args.out}  |  clean fixed-tilt (NASA POWER): "
          f"median {nasa_clean.get('median_deviation_pct')}%  "
          f"within±10% {nasa_clean.get('pct_within_10')}%")


if __name__ == "__main__":
    main()
